import openpyxl

workbook = openpyxl.Workbook()

# Seleccionar la hoja activa
hoja = workbook.active
hoja.title = "Lista de Supermercados"

# Agregar encabezados
hoja["A1"] = "Supermercado"
hoja["B1"] = "Producto"
hoja["C1"] = "Precio"

# Lista de supermercados con nombre de producto y precio
Supermercados = [
    ("Supermercado A", "Manzanas", 2.99),
    ("Supermercado A", "Peras", 3.49),
    ("Supermercado A", "Papaya", 12.12),
    ("Supermercado B", "Naranjas", 2.79),
    ("Supermercado B", "Kiwi", 3.29),
    ("Supermercado C", "Banana", 2,90),
    ("Supermercado C", "Pera", 4.18),
    
]

# Agregar datos a la hoja
for row, data in enumerate(Supermercados, start=2):
    hoja.cell(row=row, column=1, value=data[0])
    hoja.cell(row=row, column=2, value=data[1])
    hoja.cell(row=row, column=3, value=data[2])

# Guardar el archivo Excel
workbook.save("lista_supermercados.xlsx")

# Cerrar el archivo Excel
workbook.close()
